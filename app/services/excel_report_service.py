from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.report import CafeReportSchema, NetworkReportSchema


class ExcelReportService:
    def __init__(self) -> None:
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(parents=True, exist_ok=True)

        self.header_fill = PatternFill("solid", fgColor="1F4E78")
        self.section_fill = PatternFill("solid", fgColor="D9EAF7")
        self.subtle_fill = PatternFill("solid", fgColor="F3F6F9")
        self.good_fill = PatternFill("solid", fgColor="E2F0D9")
        self.bad_fill = PatternFill("solid", fgColor="FDE9E7")
        self.warning_fill = PatternFill("solid", fgColor="FFF2CC")

        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_font = Font(bold=True)
        self.bold_font = Font(bold=True)
        self.title_font = Font(bold=True, size=14)

        self.thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

    def _style_header_row(self, worksheet, row_number: int) -> None:
        for cell in worksheet[row_number]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = self.thin_border

    def _style_section_title(self, cell) -> None:
        cell.fill = self.section_fill
        cell.font = self.section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = self.thin_border

    def _style_title(self, worksheet, cell_ref: str, text: str) -> None:
        cell = worksheet[cell_ref]
        cell.value = text
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[cell.row].height = 22

    def _style_data_area(self, worksheet) -> None:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = self.thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _auto_width(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)

            worksheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                48,
            )

    @staticmethod
    def _freeze_top(worksheet, cell: str = "A2") -> None:
        worksheet.freeze_panes = cell

    def _style_kpi_table(self, worksheet, start_row: int, end_row: int) -> None:
        for row in range(start_row, end_row + 1):
            worksheet.cell(row=row, column=1).font = self.bold_font
            worksheet.cell(row=row, column=1).fill = self.subtle_fill

    @staticmethod
    def _share_percent(count: int, total: int) -> float:
        if total <= 0:
            return 0.0
        return round(count / total * 100, 2)

    @staticmethod
    def _per_100(base_count: int, total: int) -> float:
        if total <= 0:
            return 0.0
        return round(base_count / total * 100, 2)

    @staticmethod
    def _format_period(start_date, end_date) -> str:
        return f"{start_date:%d.%m.%Y} - {end_date:%d.%m.%Y}"

    @staticmethod
    def _safe_percent_change(current: int, previous: int) -> float | None:
        if previous == 0:
            return None
        return round((current - previous) / previous * 100, 2)

    def _build_sample_note(self, current_total: int, previous_total: int) -> str:
        if current_total == 0 and previous_total == 0:
            return "В обоих периодах нет анкет."
        if current_total == 0:
            return "В текущем периоде нет анкет. Сравнение с предыдущим периодом ограничено."
        if previous_total == 0:
            return "В предыдущем периоде нет анкет. Сравнение по динамике ограничено."
        if current_total < 30 or previous_total < 30:
            return "Один из периодов имеет малую выборку. Интерпретируйте динамику аккуратно."

        ratio = max(current_total, previous_total) / min(current_total, previous_total)
        if ratio >= 2:
            return "Объём выборки между периодами заметно отличается. Интерпретируйте сравнение с осторожностью."

        return "Объём выборки по периодам сопоставим для базового сравнения."

    def _apply_delta_fill(self, cell, value: float) -> None:
        if value > 0:
            cell.fill = self.good_fill
        elif value < 0:
            cell.fill = self.bad_fill

    def _append_key_value_block(
        self,
        worksheet,
        rows: list[tuple[str, object]],
    ) -> None:
        for key, value in rows:
            worksheet.append([key, value])

    def _add_focus_sheet(self, workbook: Workbook, report: NetworkReportSchema) -> None:
        ws = workbook.create_sheet(title="Фокус внимания")

        self._style_title(ws, "A1", "Фокус внимания по сети")

        ws.append([])
        ws.append(["Блок", "Кафе", "Анкет", "Средний %", "Δ п.п.", "Комментариев"])
        self._style_header_row(ws, 3)

        top_cafes = sorted(
            report.cafes,
            key=lambda item: (item.average_percent, item.total_surveys),
            reverse=True,
        )[:5]

        low_cafes = sorted(
            report.cafes,
            key=lambda item: (item.average_percent, -item.total_surveys),
        )[:5]

        falling_cafes = sorted(
            report.cafes,
            key=lambda item: item.delta_percent_points,
        )[:5]

        comment_heavy_cafes = sorted(
            report.cafes,
            key=lambda item: (item.comments_count, item.total_surveys),
            reverse=True,
        )[:5]

        def append_block(block_name: str, cafes) -> None:
            if not cafes:
                ws.append([block_name, "Нет данных", "", "", "", ""])
                return

            for cafe in cafes:
                ws.append(
                    [
                        block_name,
                        cafe.cafe_name,
                        cafe.total_surveys,
                        cafe.average_percent,
                        cafe.delta_percent_points,
                        cafe.comments_count,
                    ]
                )

        append_block("Топ-5 по среднему %", top_cafes)
        append_block("Антитоп-5 по среднему %", low_cafes)
        append_block("Наибольшая просадка", falling_cafes)
        append_block("Больше всего комментариев", comment_heavy_cafes)

        for row in range(4, ws.max_row + 1):
            delta_cell = ws.cell(row=row, column=5)
            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        self._freeze_top(ws, "A4")

    def build_cafe_report_file(self, report: CafeReportSchema) -> str:
        workbook = Workbook()

        summary_ws = workbook.active
        summary_ws.title = "Сводка"

        self._style_title(summary_ws, "A1", f"Отчёт по кафе #{report.cafe_id}")
        summary_ws.append([])

        summary_ws.append(["Показатель", "Значение"])
        self._style_header_row(summary_ws, 3)

        comparison = report.comparison
        previous_period_text = "Нет данных"
        previous_total_surveys = 0
        previous_comments_count = 0
        current_average_percent = report.summary.average_percent
        previous_average_percent = 0.0
        delta_percent_points = 0.0

        if comparison is not None:
            previous_period_text = self._format_period(
                comparison.previous_period_start_date,
                comparison.previous_period_end_date,
            )
            previous_total_surveys = comparison.previous_total_surveys
            previous_comments_count = comparison.previous_comments_count
            current_average_percent = comparison.current_average_percent
            previous_average_percent = comparison.previous_average_percent
            delta_percent_points = comparison.delta_percent_points

        current_comments_per_100 = self._per_100(
            report.comments_count,
            report.summary.total_surveys,
        )
        previous_comments_per_100 = self._per_100(
            previous_comments_count,
            previous_total_surveys,
        )
        volume_change_percent = self._safe_percent_change(
            report.summary.total_surveys,
            previous_total_surveys,
        )
        sample_note = self._build_sample_note(
            report.summary.total_surveys,
            previous_total_surveys,
        )

        self._append_key_value_block(
            summary_ws,
            [
                (
                    "Текущий период",
                    self._format_period(
                        report.period.start_date, report.period.end_date
                    ),
                ),
                ("Сопоставимый предыдущий период", previous_period_text),
                ("Анкет в текущем периоде", report.summary.total_surveys),
                ("Анкет в предыдущем периоде", previous_total_surveys),
                ("Средний балл", report.summary.average_score),
                ("Средний процент — текущий", current_average_percent),
                ("Средний процент — предыдущий", previous_average_percent),
                ("Разница, п.п.", delta_percent_points),
                ("Комментариев в текущем периоде", report.comments_count),
                ("Комментариев в предыдущем периоде", previous_comments_count),
                (
                    "Комментариев на 100 анкет — текущий период",
                    current_comments_per_100,
                ),
                (
                    "Комментариев на 100 анкет — предыдущий период",
                    previous_comments_per_100,
                ),
                (
                    "Изменение объёма выборки, %",
                    (
                        volume_change_percent
                        if volume_change_percent is not None
                        else "н/д"
                    ),
                ),
                ("Комментарий по сопоставимости", sample_note),
            ],
        )

        self._style_kpi_table(summary_ws, 4, summary_ws.max_row)

        if comparison is not None:
            delta_cell = summary_ws.cell(row=11, column=2)
            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        note_row = summary_ws.max_row
        summary_ws.cell(row=note_row, column=2).fill = self.warning_fill

        questions_ws = workbook.create_sheet(title="Вопросы")
        self._style_title(questions_ws, "A1", "Результаты по вопросам")
        questions_ws.append([])

        questions_ws.append(
            ["Вопрос", "Да", "Нет", "Да %", "Да % (пред. период)", "Разница, п.п."]
        )
        self._style_header_row(questions_ws, 3)

        stats_map = {
            "q1": report.q1_stats,
            "q2": report.q2_stats,
            "q3": report.q3_stats,
            "q4": report.q4_stats,
        }
        comparison_map = {
            item.question_key: item for item in report.question_comparisons
        }

        for question_key, stats in stats_map.items():
            comparison_item = comparison_map.get(question_key)
            questions_ws.append(
                [
                    comparison_item.question_label if comparison_item else question_key,
                    stats.yes_count,
                    stats.no_count,
                    stats.yes_percent,
                    comparison_item.previous_yes_percent if comparison_item else 0.0,
                    comparison_item.delta_percent_points if comparison_item else 0.0,
                ]
            )

        for row in range(4, questions_ws.max_row + 1):
            delta_cell = questions_ws.cell(row=row, column=6)
            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        distribution_ws = workbook.create_sheet(title="Распределение")
        self._style_title(distribution_ws, "A1", "Распределение оценок")
        distribution_ws.append([])

        distribution_ws.append(
            [
                "Оценка",
                "Текущий период, шт.",
                "Текущий период, %",
                "Предыдущий период, шт.",
                "Предыдущий период, %",
                "Изменение доли, п.п.",
            ]
        )
        self._style_header_row(distribution_ws, 3)

        current_total = report.summary.total_surveys
        previous_total = previous_total_surveys

        for item in report.distribution_comparisons:
            current_share = self._share_percent(item.current_count, current_total)
            previous_share = self._share_percent(item.previous_count, previous_total)
            distribution_ws.append(
                [
                    item.label,
                    item.current_count,
                    current_share,
                    item.previous_count,
                    previous_share,
                    round(current_share - previous_share, 2),
                ]
            )

        for row in range(4, distribution_ws.max_row + 1):
            score_label = distribution_ws.cell(row=row, column=1).value
            delta_cell = distribution_ws.cell(row=row, column=6)

            if score_label == "100%":
                for col in range(1, 7):
                    distribution_ws.cell(row=row, column=col).fill = self.good_fill
            elif score_label == "0%":
                for col in range(1, 7):
                    distribution_ws.cell(row=row, column=col).fill = self.bad_fill

            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        dynamics_ws = workbook.create_sheet(title="Динамика")
        self._style_title(dynamics_ws, "A1", "Динамика по периодам")
        dynamics_ws.append([])

        dynamics_ws.append(["Недельная динамика", "", "", "", "", ""])
        self._style_section_title(dynamics_ws["A3"])
        dynamics_ws.append(
            [
                "Период",
                "Анкет",
                "Средний балл",
                "Средний %",
                "Комментариев",
                "Комментариев на 100 анкет",
            ]
        )
        self._style_header_row(dynamics_ws, 4)

        if report.dynamics is not None:
            for point in report.dynamics.weekly_points:
                average_score = round(point.average_percent / 25, 2)
                dynamics_ws.append(
                    [
                        point.label,
                        point.total_surveys,
                        average_score,
                        point.average_percent,
                        point.comments_count,
                        self._per_100(point.comments_count, point.total_surveys),
                    ]
                )

        monthly_start_row = dynamics_ws.max_row + 2
        dynamics_ws.cell(row=monthly_start_row, column=1, value="Месячная динамика")
        self._style_section_title(dynamics_ws.cell(row=monthly_start_row, column=1))

        dynamics_ws.append(
            [
                "Период",
                "Анкет",
                "Средний балл",
                "Средний %",
                "Комментариев",
                "Комментариев на 100 анкет",
            ]
        )
        self._style_header_row(dynamics_ws, monthly_start_row + 1)

        if report.dynamics is not None:
            for point in report.dynamics.monthly_points:
                average_score = round(point.average_percent / 25, 2)
                dynamics_ws.append(
                    [
                        point.label,
                        point.total_surveys,
                        average_score,
                        point.average_percent,
                        point.comments_count,
                        self._per_100(point.comments_count, point.total_surveys),
                    ]
                )

        comments_ws = workbook.create_sheet(title="Комментарии")
        self._style_title(comments_ws, "A1", "Комментарии гостей")
        comments_ws.append([])

        comments_ws.append(["№", "Комментарий"])
        self._style_header_row(comments_ws, 3)

        if report.comments:
            for index, comment in enumerate(report.comments, start=1):
                comments_ws.append([index, comment])
        else:
            comments_ws.append([1, "Нет комментариев за выбранный период"])

        ai_ws = workbook.create_sheet(title="AI-анализ")
        self._style_title(ai_ws, "A1", "AI-анализ комментариев")
        ai_ws.append([])

        ai_ws.append(["AI-анализ"])
        self._style_header_row(ai_ws, 3)

        ai_lines = (
            report.ai_summary.splitlines()
            if report.ai_summary
            else ["AI-анализ недоступен."]
        )
        for line in ai_lines:
            ai_ws.append([line])

        for ws in workbook.worksheets:
            self._style_data_area(ws)
            self._auto_width(ws)

        self._freeze_top(summary_ws, "A4")
        self._freeze_top(questions_ws, "A4")
        self._freeze_top(distribution_ws, "A4")
        self._freeze_top(dynamics_ws, "A5")
        self._freeze_top(comments_ws, "A4")
        self._freeze_top(ai_ws, "A4")

        file_name = (
            f"cafe_report_{report.cafe_id}_"
            f"{report.period.start_date:%Y%m%d}_{report.period.end_date:%Y%m%d}.xlsx"
        )
        file_path = self.reports_dir / file_name
        workbook.save(file_path)

        return str(file_path)

    def build_network_report_file(self, report: NetworkReportSchema) -> str:
        workbook = Workbook()

        summary_ws = workbook.active
        summary_ws.title = "Сводка"

        self._style_title(summary_ws, "A1", "Общий отчёт по сети")
        summary_ws.append([])

        summary_ws.append(["Показатель", "Значение"])
        self._style_header_row(summary_ws, 3)

        previous_total = sum(
            item.previous_count for item in report.distribution_comparisons
        )
        current_comments_per_100 = self._per_100(
            report.comments_count, report.total_surveys
        )
        previous_comments_per_100 = self._per_100(
            report.previous_comments_count,
            previous_total,
        )
        sample_note = self._build_sample_note(report.total_surveys, previous_total)
        volume_change_percent = self._safe_percent_change(
            report.total_surveys, previous_total
        )

        self._append_key_value_block(
            summary_ws,
            [
                (
                    "Текущий период",
                    self._format_period(
                        report.period.start_date, report.period.end_date
                    ),
                ),
                ("Кафе в отчёте", report.total_cafes),
                ("Анкет в текущем периоде", report.total_surveys),
                ("Анкет в предыдущем периоде", previous_total),
                ("Средний балл", report.average_score),
                ("Средний процент по сети", report.average_percent),
                ("Комментариев в текущем периоде", report.comments_count),
                ("Комментариев в предыдущем периоде", report.previous_comments_count),
                (
                    "Комментариев на 100 анкет — текущий период",
                    current_comments_per_100,
                ),
                (
                    "Комментариев на 100 анкет — предыдущий период",
                    previous_comments_per_100,
                ),
                (
                    "Изменение объёма выборки, %",
                    (
                        volume_change_percent
                        if volume_change_percent is not None
                        else "н/д"
                    ),
                ),
                (
                    "Логика сравнения",
                    "Сравнение идёт с предыдущим периодом такой же длины, непосредственно предшествующим текущему.",
                ),
                ("Комментарий по сопоставимости", sample_note),
            ],
        )

        self._style_kpi_table(summary_ws, 4, summary_ws.max_row)
        summary_ws.cell(row=summary_ws.max_row, column=2).fill = self.warning_fill

        questions_ws = workbook.create_sheet(title="Вопросы")
        self._style_title(questions_ws, "A1", "Результаты по вопросам")
        questions_ws.append([])

        questions_ws.append(
            ["Вопрос", "Да", "Нет", "Да %", "Да % (пред. период)", "Разница, п.п."]
        )
        self._style_header_row(questions_ws, 3)

        stats_map = {
            "q1": report.q1_stats,
            "q2": report.q2_stats,
            "q3": report.q3_stats,
            "q4": report.q4_stats,
        }
        comparison_map = {
            item.question_key: item for item in report.question_comparisons
        }

        for question_key, stats in stats_map.items():
            comparison_item = comparison_map.get(question_key)
            questions_ws.append(
                [
                    comparison_item.question_label if comparison_item else question_key,
                    stats.yes_count,
                    stats.no_count,
                    stats.yes_percent,
                    comparison_item.previous_yes_percent if comparison_item else 0.0,
                    comparison_item.delta_percent_points if comparison_item else 0.0,
                ]
            )

        for row in range(4, questions_ws.max_row + 1):
            delta_cell = questions_ws.cell(row=row, column=6)
            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        distribution_ws = workbook.create_sheet(title="Распределение")
        self._style_title(distribution_ws, "A1", "Распределение оценок по сети")
        distribution_ws.append([])

        distribution_ws.append(
            [
                "Оценка",
                "Текущий период, шт.",
                "Текущий период, %",
                "Предыдущий период, шт.",
                "Предыдущий период, %",
                "Изменение доли, п.п.",
            ]
        )
        self._style_header_row(distribution_ws, 3)

        current_total = report.total_surveys

        for item in report.distribution_comparisons:
            current_share = self._share_percent(item.current_count, current_total)
            previous_share = self._share_percent(item.previous_count, previous_total)
            distribution_ws.append(
                [
                    item.label,
                    item.current_count,
                    current_share,
                    item.previous_count,
                    previous_share,
                    round(current_share - previous_share, 2),
                ]
            )

        for row in range(4, distribution_ws.max_row + 1):
            score_label = distribution_ws.cell(row=row, column=1).value
            delta_cell = distribution_ws.cell(row=row, column=6)

            if score_label == "100%":
                for col in range(1, 7):
                    distribution_ws.cell(row=row, column=col).fill = self.good_fill
            elif score_label == "0%":
                for col in range(1, 7):
                    distribution_ws.cell(row=row, column=col).fill = self.bad_fill

            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        dynamics_ws = workbook.create_sheet(title="Динамика")
        self._style_title(dynamics_ws, "A1", "Динамика по сети")
        dynamics_ws.append([])

        dynamics_ws.append(["Недельная динамика сети", "", "", "", "", ""])
        self._style_section_title(dynamics_ws["A3"])
        dynamics_ws.append(
            [
                "Период",
                "Анкет",
                "Средний балл",
                "Средний %",
                "Комментариев",
                "Комментариев на 100 анкет",
            ]
        )
        self._style_header_row(dynamics_ws, 4)

        if report.dynamics is not None:
            for point in report.dynamics.weekly_points:
                average_score = round(point.average_percent / 25, 2)
                dynamics_ws.append(
                    [
                        point.label,
                        point.total_surveys,
                        average_score,
                        point.average_percent,
                        point.comments_count,
                        self._per_100(point.comments_count, point.total_surveys),
                    ]
                )

        monthly_start_row = dynamics_ws.max_row + 2
        dynamics_ws.cell(
            row=monthly_start_row, column=1, value="Месячная динамика сети"
        )
        self._style_section_title(dynamics_ws.cell(row=monthly_start_row, column=1))

        dynamics_ws.append(
            [
                "Период",
                "Анкет",
                "Средний балл",
                "Средний %",
                "Комментариев",
                "Комментариев на 100 анкет",
            ]
        )
        self._style_header_row(dynamics_ws, monthly_start_row + 1)

        if report.dynamics is not None:
            for point in report.dynamics.monthly_points:
                average_score = round(point.average_percent / 25, 2)
                dynamics_ws.append(
                    [
                        point.label,
                        point.total_surveys,
                        average_score,
                        point.average_percent,
                        point.comments_count,
                        self._per_100(point.comments_count, point.total_surveys),
                    ]
                )

        cafes_ws = workbook.create_sheet(title="Кафе")
        self._style_title(cafes_ws, "A1", "Показатели по кафе")
        cafes_ws.append([])

        cafes_ws.append(
            [
                "ID кафе",
                "Название кафе",
                "Анкет",
                "Средний %",
                "Средний % (пред. период)",
                "Разница, п.п.",
                "Комментариев",
                "Комментариев на 100 анкет",
            ]
        )
        self._style_header_row(cafes_ws, 3)

        for cafe in report.cafes:
            cafes_ws.append(
                [
                    cafe.cafe_id,
                    cafe.cafe_name,
                    cafe.total_surveys,
                    cafe.average_percent,
                    cafe.previous_average_percent,
                    cafe.delta_percent_points,
                    cafe.comments_count,
                    self._per_100(cafe.comments_count, cafe.total_surveys),
                ]
            )

        for row in range(4, cafes_ws.max_row + 1):
            delta_cell = cafes_ws.cell(row=row, column=6)
            if isinstance(delta_cell.value, (int, float)):
                self._apply_delta_fill(delta_cell, float(delta_cell.value))

        self._add_focus_sheet(workbook, report)

        ai_ws = workbook.create_sheet(title="AI-анализ")
        self._style_title(ai_ws, "A1", "AI-анализ по сети")
        ai_ws.append([])

        ai_ws.append(["AI-анализ"])
        self._style_header_row(ai_ws, 3)

        ai_lines = (
            report.ai_summary.splitlines()
            if report.ai_summary
            else ["AI-анализ недоступен."]
        )
        for line in ai_lines:
            ai_ws.append([line])

        for ws in workbook.worksheets:
            self._style_data_area(ws)
            self._auto_width(ws)

        self._freeze_top(summary_ws, "A4")
        self._freeze_top(questions_ws, "A4")
        self._freeze_top(distribution_ws, "A4")
        self._freeze_top(dynamics_ws, "A5")
        self._freeze_top(cafes_ws, "A4")
        self._freeze_top(ai_ws, "A4")

        file_name = (
            f"network_report_"
            f"{report.period.start_date:%Y%m%d}_{report.period.end_date:%Y%m%d}.xlsx"
        )
        file_path = self.reports_dir / file_name
        workbook.save(file_path)

        return str(file_path)
